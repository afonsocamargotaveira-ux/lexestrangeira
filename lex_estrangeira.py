import streamlit as st
import pandas as pd
from urllib.parse import quote
from datetime import datetime
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="LexEstrangeira",
    page_icon="=",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── CSS ───────────────────────────────────────────────────────────────────────
CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;700&family=DM+Sans:wght@300;400;500;600&display=swap');

:root {
    --bg:       #1a1d23;
    --surface:  #22262f;
    --surface2: #2b3040;
    --border:   #353a47;
    --accent:   #c9a84c;
    --accent2:  #e8c97a;
    --text:     #e8eaf0;
    --muted:    #8b93a8;
    --radius:   10px;
}

html, body, [data-testid="stAppViewContainer"] {
    background-color: var(--bg) !important;
    color: var(--text) !important;
    font-family: 'DM Sans', sans-serif !important;
}
[data-testid="stHeader"] { display: none !important; }
[data-testid="stToolbar"] { display: none !important; }
footer { display: none !important; }
#MainMenu { display: none !important; }
.block-container { padding-top: 1.5rem !important; max-width: 1100px; }

.logo-wrapper {
    display: flex; align-items: center; gap: 14px;
    padding: 0.5rem 0 1.4rem 0;
}
.logo-icon {
    width: 48px; height: 48px;
    background: linear-gradient(135deg, #c9a84c, #e8c97a);
    border-radius: 11px;
    display: flex; align-items: center; justify-content: center;
    font-size: 24px;
    box-shadow: 0 4px 18px rgba(201,168,76,0.38);
    flex-shrink: 0;
}
.logo-title {
    font-family: 'Playfair Display', serif;
    font-size: 1.85rem; font-weight: 700;
    color: var(--text); letter-spacing: -0.02em; line-height: 1;
}
.logo-title span { color: var(--accent); }
.logo-sub {
    font-size: 0.7rem; color: var(--muted);
    letter-spacing: 0.13em; text-transform: uppercase; margin-top: 4px;
}
.divider {
    border: none; border-top: 1px solid var(--border); margin: 0 0 1.5rem 0;
}
.search-card {
    background: var(--surface); border: 1px solid var(--border);
    border-radius: 14px; padding: 1.5rem 1.75rem; margin-bottom: 1.5rem;
}
.search-card h3 {
    font-family: 'Playfair Display', serif; font-size: 1rem;
    color: var(--accent); margin-bottom: 1rem; font-weight: 700;
}
.stTextInput > div > div > input,
.stSelectbox > div > div {
    background: var(--surface2) !important;
    border: 1px solid var(--border) !important;
    color: var(--text) !important;
    border-radius: var(--radius) !important;
    font-family: 'DM Sans', sans-serif !important;
}
.stTextInput > div > div > input:focus {
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 2px rgba(201,168,76,0.18) !important;
}
label {
    color: var(--muted) !important; font-size: 0.76rem !important;
    letter-spacing: 0.07em !important; text-transform: uppercase !important;
    font-weight: 500 !important;
}
.stButton > button {
    background: linear-gradient(135deg, var(--accent), var(--accent2)) !important;
    color: #111 !important; font-weight: 600 !important;
    font-family: 'DM Sans', sans-serif !important;
    border: none !important; border-radius: var(--radius) !important;
    padding: 0.55rem 1.5rem !important;
    transition: opacity 0.2s, transform 0.15s !important;
}
.stButton > button:hover {
    opacity: 0.85 !important; transform: translateY(-1px) !important;
}
.stDownloadButton > button {
    background: var(--surface2) !important; color: var(--accent2) !important;
    border: 1px solid var(--accent) !important; font-weight: 500 !important;
    font-family: 'DM Sans', sans-serif !important;
    border-radius: var(--radius) !important;
}
.stDownloadButton > button:hover {
    background: rgba(201,168,76,0.1) !important;
}
.stats-row {
    display: flex; gap: 0.85rem; margin: 1rem 0 1.3rem 0; flex-wrap: wrap;
}
.stat-pill {
    background: var(--surface2); border: 1px solid var(--border);
    border-radius: 8px; padding: 0.4rem 0.85rem;
    font-size: 0.79rem; color: var(--muted);
}
.stat-pill b { color: var(--accent2); }
</style>
"""
st.markdown(CSS, unsafe_allow_html=True)

# ── Dados dos paises ──────────────────────────────────────────────────────────
COUNTRIES = [
    {
        "pais": "Australia",
        "flag": "AU",
        "template": (
            "https://www.legislation.gov.au/search/"
            "text(%22{term}%22,nameAndText,contains)/pointintime(Latest)"
        ),
    },
    {
        "pais": "Nova Zelandia",
        "flag": "NZ",
        "template": (
            "https://www.legislation.govt.nz/items/"
            "?search_term={term}&search_field=title&per_page=20"
            "&search_for=in_force&legislation_status=in_force"
        ),
    },
    {
        "pais": "Reino Unido",
        "flag": "UK",
        "template": "https://www.legislation.gov.uk/all?text={term}",
    },
    {
        "pais": "Canada",
        "flag": "CA",
        "template": "https://www.parl.ca/LegisInfo/en/bills?keywords={term}",
    },
    {
        "pais": "Irlanda",
        "flag": "IE",
        "template": "https://www.irishstatutebook.ie/eli/ResultsTitle.html?q={term}",
    },
    {
        "pais": "India",
        "flag": "IN",
        "template": (
            "https://www.indiacode.nic.in/handle/123456789/1362/simple-search"
            "?page-token=11dab4b4c061&page-token-value=11b9aeff354815694cb44ab87c5d4439"
            "&nccharset=DCE5AB6E&query={term}&btngo=&searchradio=acts"
        ),
    },
    {
        "pais": "Estados Unidos",
        "flag": "US",
        "template": (
            "https://www.govinfo.gov/app/search/"
            "%7B%22historical%22%3Atrue%2C%22offset%22%3A0%2C%22query%22%3A"
            "%22collection%3A(GPO%20OR%20BUDGET%20OR%20CZIC%20OR%20CFR%20OR%20CPD"
            "%20OR%20BILLS%20OR%20CCAL%20OR%20CPRT%20OR%20CDIR%20OR%20CDOC"
            "%20OR%20CHRG%20OR%20CREC%20OR%20CRECB%20OR%20CRI%20OR%20CRPT"
            "%20OR%20SERIALSET%20OR%20CMR%20OR%20ECONI%20OR%20ERP%20OR%20ERIC"
            "%20OR%20FR%20OR%20GAOREPORTS%20OR%20HOB%20OR%20HMAN%20OR%20HJOURNAL"
            "%20OR%20SJOURNAL%20OR%20LSA%20OR%20GOVPUB%20OR%20PAI%20OR%20PPP"
            "%20OR%20PLAW%20OR%20SMAN%20OR%20COMPS%20OR%20STATUTE%20OR%20USCODE"
            "%20OR%20USCOURTS%20OR%20GOVMAN%20OR%20USREPORTS)"
            "%20AND%20publishdate%3Arange(%2C2026-06-01)"
            "%20%20AND%20%20content%3A({term})%22%7D"
        ),
    },
]


def gerar_links(term, paises_selecionados):
    encoded = quote(term)
    rows = []
    for c in COUNTRIES:
        if c["pais"] not in paises_selecionados:
            continue
        url = c["template"].replace("{term}", encoded)
        rows.append({"pais": c["pais"], "flag": c["flag"], "link": url})
    return rows


def exportar_excel(rows, term):
    wb = Workbook()
    ws = wb.active
    ws.title = "Links de Busca"

    header_fill = PatternFill("solid", fgColor="22262F")
    accent_font = Font(bold=True, color="C9A84C", name="Calibri", size=11)
    link_font   = Font(color="4FC3F7", name="Calibri", size=11, underline="single")
    row_fill    = PatternFill("solid", fgColor="22262F")
    alt_fill    = PatternFill("solid", fgColor="2B3040")
    thin_border = Border(bottom=Side(style="thin", color="353A47"))
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:B1")
    ws["A1"] = "LexEstrangeira -- Busca: " + term
    ws["A1"].font = Font(bold=True, color="C9A84C", name="Calibri", size=14)
    ws["A1"].alignment = center
    ws["A1"].fill = PatternFill("solid", fgColor="1A1D23")
    ws.row_dimensions[1].height = 30

    ws.append([""])
    ws.append(["Pais", "Link de Busca Oficial"])

    hdr_row = 3
    for col, h in enumerate(["Pais", "Link de Busca Oficial"], 1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.font = accent_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[hdr_row].height = 22

    for i, r in enumerate(rows):
        data_row = hdr_row + 1 + i
        fill = row_fill if i % 2 == 0 else alt_fill

        pais_cell = ws.cell(row=data_row, column=1, value=r["flag"] + "  " + r["pais"])
        pais_cell.font = Font(bold=True, color="E8EAF0", name="Calibri", size=11)
        pais_cell.fill = fill
        pais_cell.alignment = center
        pais_cell.border = thin_border

        link_cell = ws.cell(row=data_row, column=2, value=r["link"])
        link_cell.font = link_font
        link_cell.fill = fill
        link_cell.alignment = left
        link_cell.border = thin_border
        link_cell.hyperlink = r["link"]

        ws.row_dimensions[data_row].height = 20

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── UI ────────────────────────────────────────────────────────────────────────

st.markdown(
    '<div class="logo-wrapper">'
    '<div class="logo-icon">&#9878;</div>'
    '<div style="line-height:1">'
    '<div class="logo-title">Lex<span>Estrangeira</span></div>'
    '<div class="logo-sub">Busca de Legislacao Internacional</div>'
    '</div></div>'
    '<hr class="divider">',
    unsafe_allow_html=True,
)

st.markdown('<div class="search-card"><h3>&#128269; Parametros de Busca</h3>', unsafe_allow_html=True)

col1, col2 = st.columns([2, 3])
with col1:
    term = st.text_input("Termo de Busca", placeholder="Ex: police, environment, education...")
with col2:
    all_paises = [c["pais"] for c in COUNTRIES]
    opcoes = ["Todos os paises"] + all_paises
    selecionado = st.selectbox("Pais", opcoes)

col_btn, _ = st.columns([1, 5])
with col_btn:
    buscar = st.button("Gerar Links", use_container_width=True)

st.markdown("</div>", unsafe_allow_html=True)

# ── Resultado ─────────────────────────────────────────────────────────────────
if buscar:
    if not term.strip():
        st.warning("Por favor, digite um termo de busca.")
    else:
        paises_alvo = all_paises if selecionado == "Todos os paises" else [selecionado]
        rows = gerar_links(term.strip(), paises_alvo)

        st.markdown(
            '<div class="stats-row">'
            '<div class="stat-pill">Termo: <b>' + term.strip() + '</b></div>'
            '<div class="stat-pill">Paises: <b>' + str(len(rows)) + '</b></div>'
            '</div>',
            unsafe_allow_html=True,
        )

        excel_bytes = exportar_excel(rows, term.strip())
        st.download_button(
            label="Exportar para Excel",
            data=excel_bytes,
            file_name=(
                "lex_estrangeira_"
                + term.strip().replace(" ", "_")
                + "_"
                + datetime.today().strftime("%Y%m%d")
                + ".xlsx"
            ),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Cada card tem um <a href> puro — sem passar por componente Streamlit
        # para evitar truncamento de URLs longas
        parts = ['<div style="margin-top:1rem">']
        for r in rows:
            card = (
                '<div style="'
                "background:#22262f;"
                "border:1px solid #353a47;"
                "border-left:4px solid #c9a84c;"
                "border-radius:10px;"
                "padding:0.85rem 1.1rem;"
                "margin-bottom:0.75rem;"
                "display:flex;"
                "align-items:center;"
                "justify-content:space-between;"
                'gap:1rem">'
                '<div style="flex:1;min-width:0">'
                '<span style="font-weight:600;font-size:1rem;color:#e8eaf0">'
                + r["flag"] + "  " + r["pais"] +
                "</span><br>"
                '<span style="font-size:0.73rem;color:#8b93a8;word-break:break-all">'
                + r["link"] +
                "</span>"
                "</div>"
                '<a href="' + r["link"] + '" target="_blank" rel="noopener noreferrer" '
                'style="'
                "flex-shrink:0;"
                "background:#2b3040;"
                "color:#e8c97a;"
                "border:1px solid #c9a84c;"
                "border-radius:8px;"
                "padding:0.45rem 1rem;"
                "font-size:0.85rem;"
                "font-weight:500;"
                "text-decoration:none;"
                'white-space:nowrap">'
                "Abrir"
                "</a>"
                "</div>"
            )
            parts.append(card)
        parts.append("</div>")
        st.markdown("".join(parts), unsafe_allow_html=True)

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown(
    '<hr class="divider" style="margin-top:3rem">'
    '<div style="text-align:center;color:#8b93a8;font-size:0.73rem;padding-bottom:1rem">'
    '<b style="color:#c9a84c">LexEstrangeira</b>'
    " &nbsp;&middot;&nbsp; Australia &nbsp;&middot;&nbsp; Nova Zelandia"
    " &nbsp;&middot;&nbsp; Reino Unido &nbsp;&middot;&nbsp; Canada"
    " &nbsp;&middot;&nbsp; Estados Unidos"
    "</div>",
    unsafe_allow_html=True,
)
